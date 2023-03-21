from openpyxl import load_workbook
import re
workbook = load_workbook(filename="info-sheet.xlsx")
workbook.sheetnames

sheet = workbook.active
countriesList = ['NO DATA', 'Unknown', 'UK', 'USA', 'Afghanistan', 'Aland Islands', 'Albania', 'Algeria', 'American Samoa', 'Andorra', 'Angola', 'Anguilla', 'Antarctica', 'Antigua and Barbuda', 'Argentina', 'Armenia', 'Aruba', 'Australia', 'Austria', 'Azerbaijan', 'Bahamas', 'Bahrain', 'Bangladesh', 'Barbados', 'Belarus', 'Belgium', 'Belize', 'Benin', 'Bermuda', 'Bhutan', 'Bolivia, Plurinational State of', 'Bonaire, Sint Eustatius and Saba', 'Bosnia and Herzegovina', 'Botswana', 'Bouvet Island', 'Brazil', 'British Indian Ocean Territory', 'Brunei Darussalam', 'Bulgaria', 'Burkina Faso', 'Burundi', 'Cambodia', 'Cameroon', 'Canada', 'Cape Verde', 'Cayman Islands', 'Central African Republic', 'Chad', 'Chile', 'China', 'Christmas Island', 'Cocos (Keeling) Islands', 'Colombia', 'Comoros', 'Congo', 'Congo, The Democratic Republic of the', 'Cook Islands', 'Costa Rica', "Côte d'Ivoire", 'Croatia', 'Cuba', 'Curaçao', 'Cyprus', 'Czech Republic', 'Denmark', 'Djibouti', 'Dominica', 'Dominican Republic', 'Ecuador', 'Egypt', 'El Salvador', 'Equatorial Guinea', 'Eritrea', 'Estonia', 'Ethiopia', 'Falkland Islands (Malvinas)', 'Faroe Islands', 'Fiji', 'Finland', 'France', 'French Guiana', 'French Polynesia', 'French Southern Territories', 'Gabon', 'Gambia', 'Georgia', 'Germany', 'Ghana', 'Gibraltar', 'Greece', 'Greenland', 'Grenada', 'Guadeloupe', 'Guam', 'Guatemala', 'Guernsey', 'Guinea', 'Guinea-Bissau', 'Guyana', 'Haiti', 'Heard Island and McDonald Islands', 'Holy See (Vatican City State)', 'Honduras', 'Hong Kong', 'Hungary', 'Iceland', 'India', 'Indonesia', 'Iran, Islamic Republic of', 'Iraq', 'Ireland', 'Isle of Man', 'Israel', 'Italy', 'Jamaica', 'Japan', 'Jersey', 'Jordan', 'Kazakhstan', 'Kenya', 'Kiribati', "Korea, Democratic People's Republic of", 'Korea, Republic of', 'Kuwait', 'Kyrgyzstan', "Lao People's Democratic Republic", 'Latvia', 'Lebanon', 'Lesotho', 'Liberia', 'Libya', 'Liechtenstein', 'Lithuania', 'Luxembourg', 'Macao', 'Macedonia, Republic of', 'Madagascar', 'Malawi', 'Malaysia', 'Maldives', 'Mali', 'Malta', 'Marshall Islands', 'Martinique', 'Mauritania', 'Mauritius', 'Mayotte', 'Mexico', 'Micronesia, Federated States of', 'Moldova, Republic of', 'Monaco', 'Mongolia', 'Montenegro', 'Montserrat', 'Morocco', 'Mozambique', 'Myanmar', 'Namibia', 'Nauru', 'Nepal', 'Netherlands', 'New Caledonia', 'New Zealand', 'Nicaragua', 'Niger', 'Nigeria', 'Niue', 'Norfolk Island', 'Northern Mariana Islands', 'Norway', 'Oman', 'Pakistan', 'Palau', 'Palestinian Territory, Occupied', 'Panama', 'Papua New Guinea', 'Paraguay', 'Peru', 'Philippines', 'Pitcairn', 'Poland', 'Portugal', 'Puerto Rico', 'Qatar', 'Réunion', 'Romania', 'Russian Federation', 'Rwanda', 'Saint Barthélemy', 'Saint Helena, Ascension and Tristan da Cunha', 'Saint Kitts and Nevis', 'Saint Lucia', 'Saint Martin (French part)', 'Saint Pierre and Miquelon', 'Saint Vincent and the Grenadines', 'Samoa', 'San Marino', 'Sao Tome and Principe', 'Saudi Arabia', 'Senegal', 'Serbia', 'Seychelles', 'Sierra Leone', 'Singapore', 'Sint Maarten (Dutch part)', 'Slovakia', 'Slovenia', 'Solomon Islands', 'Somalia', 'South Africa', 'South Georgia and the South Sandwich Islands', 'Spain', 'Sri Lanka', 'Sudan', 'Suriname', 'South Sudan', 'Svalbard and Jan Mayen', 'Swaziland', 'Sweden', 'Switzerland', 'Syrian Arab Republic', 'Taiwan, Province of China', 'Tajikistan', 'Tanzania, United Republic of', 'Thailand', 'Timor-Leste', 'Togo', 'Tokelau', 'Tonga', 'Trinidad and Tobago', 'Tunisia', 'Turkey', 'Turkmenistan', 'Turks and Caicos Islands', 'Tuvalu', 'Uganda', 'Ukraine', 'United Arab Emirates', 'United States Minor Outlying Islands', 'Uruguay', 'Uzbekistan', 'Vanuatu', 'Venezuela, Bolivarian Republic of', 'Viet Nam', 'Virgin Islands, British', 'Virgin Islands, U.S.', 'Wallis and Futuna', 'Yemen', 'Zambia', 'Zimbabwe']
list = []
countries = {}

for x in range(2, 901):
    list.append(sheet['C'+str(x)].value)

for x in range(0, 899):
    if list[x]:
        split = list[x].split()
        if split[len(split)-1] == '[?]':
            list[x] = split[len(split) - 2]
        elif split[0] == 'Mexico':
            list[x] = 'Mexico'
        elif split[0] == 'Mexico;':
            list[x] = 'Mexico'
        else:
            list[x] = split[len(split)-1]
    else:
        list[x] = 'NO DATA'


def get_match(source, index):
    for pattern in countriesList:
        result = re.match(source, pattern)
        if result:
            countries[index] = pattern
            break

        result = re.match(source, "Ceylon")
        if result:
            countries[index] = "Sri Lanka"
            break

        result = re.match(source, "London")
        if result:
            countries[index] = "UK"
            break

        result = re.match(source, "England")
        if result:
            countries[index] = "UK"
            break

        result = re.match(source, "Scotland")
        if result:
            countries[index] = "UK"
            break

        result = re.match(source, "Britain")
        if result:
            countries[index] = "UK"
            break

        result = re.match(source, "U.K")
        if result:
            countries[index] = "UK"
            break

        result = re.match(source, "U.K.")
        if result:
            countries[index] = "UK"
            break

        result = re.match(source, "Zealand")
        if result:
            countries[index] = "New Zealand"
            break

        result = re.match(source, "U.S.A.")
        if result:
            countries[index] = "USA"
            break

        result = re.match(source, "States")
        if result:
            countries[index] = "USA"
            break

        result = re.match(source, "US")
        if result:
            countries[index] = "USA"
            break

        result = re.match(source, "NY")
        if result:
            countries[index] = "USA"
            break

        result = re.match(source, "York")
        if result:
            countries[index] = "USA"
            break

        result = re.match(source, "City")
        if result:
            countries[index] = "USA"
            break

        result = re.match(source, "Pennsylvania")
        if result:
            countries[index] = "USA"
            break

        result = re.match(source, "Missouri")
        if result:
            countries[index] = "USA"
            break

        result = re.match(source, "Kansas")
        if result:
            countries[index] = "USA"
            break

        result = re.match(source, "Missouri")
        if result:
            countries[index] = "USA"
            break

        countries[index] = "!NO MATCH"


for i in range(0, 899):
    src = str(list[i])
    src = src.replace(")", "")
    list[i] = src

for i in range(0, 899):
    src = str(list[i])
    get_match(src, i)

count1 = 0
count2 = 0
count3 = 0
for i in range(0, 899):
    if countries[i] == '!NO MATCH':
        count1 = count1+1
        continue

    if countries[i] == 'NO DATA':
        count3 = count3+1
        continue

    count2 = count2+1

for x in range(2, 901):
    sheet["C"+str(x)] = countries[x-2]

workbook.save(filename="parsed-info-sheet.xslx")

print("No match:")
print(count1)
print("Match:")
print(count2)
print("No data:")
print(count3)
print(countries)
